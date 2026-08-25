#!/usr/bin/env python3
"""
scripts/importar_produtos.py
---------------------------------------------------------------
Converte a planilha de exportacao do PDV (ex.: Produtos_881.xlsx) para o
formato de seed usado pelo app (data/produtos.seed.json).

Uso:
    python3 importar_produtos.py <caminho_planilha.xlsx> <caminho_saida.json>

Este script e a REFERENCIA de mapeamento de campos. A tela "Importar
Planilha" do Cadastro de Produtos (no navegador, via SheetJS) deve seguir
exatamente a mesma logica -- ver src/lib/importarProdutos.ts.

Campos fiscais do PDV (NCM, CFOP, ICMS, PIS, COFINS, etc.) sao
IGNORADOS deliberadamente: pertencem ao sistema fiscal/NF-e, fora do
escopo do app de Producao e Perdas.
"""

import json
import sys
from dataclasses import dataclass, field, asdict

try:
    import openpyxl
except ImportError:
    print("Dependencia faltando: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

COL_CATEGORIA = 2
COL_NOME = 3
COL_PRECO_CUSTO = 4
COL_PRECO_VENDA = 5
COL_MEDIDA = 6
COL_STATUS_VENDA = 11
COL_COD_PDV = 12

CATEGORIA_PADRAO = "SEM_CATEGORIA"


@dataclass
class RelatorioImportacao:
    total_linhas: int = 0
    importados: int = 0
    sem_categoria: int = 0
    codigos_pdv_duplicados: list = field(default_factory=list)
    unidades_nao_reconhecidas: list = field(default_factory=list)


def normalizar_unidade(medida_raw: str, nome_produto: str, avisos: list) -> str:
    m = (medida_raw or "").strip().lower()
    if m in ("un", "kg", "l"):
        return m
    avisos.append(f'"{nome_produto}": unidade "{medida_raw}" nao reconhecida, usando "un"')
    return "un"


def importar(caminho_entrada: str, caminho_saida: str) -> RelatorioImportacao:
    wb = openpyxl.load_workbook(caminho_entrada, data_only=True)
    ws = wb["Sheet"]

    relatorio = RelatorioImportacao()
    produtos = []
    codigos_vistos = set()

    for r in range(2, ws.max_row + 1):
        nome = ws.cell(row=r, column=COL_NOME).value
        cod_pdv = ws.cell(row=r, column=COL_COD_PDV).value
        if nome is None or cod_pdv is None:
            continue  # linha vazia / de rodape

        relatorio.total_linhas += 1

        categoria_raw = (ws.cell(row=r, column=COL_CATEGORIA).value or "").strip()
        categoria = categoria_raw if categoria_raw and categoria_raw != "..." else CATEGORIA_PADRAO
        if categoria == CATEGORIA_PADRAO:
            relatorio.sem_categoria += 1

        unidade = normalizar_unidade(
            ws.cell(row=r, column=COL_MEDIDA).value, nome, relatorio.unidades_nao_reconhecidas
        )

        cod_pdv_int = int(cod_pdv)
        if cod_pdv_int in codigos_vistos:
            relatorio.codigos_pdv_duplicados.append(cod_pdv_int)
            continue
        codigos_vistos.add(cod_pdv_int)

        produtos.append(
            {
                "codigoPdv": cod_pdv_int,
                "nome": str(nome).strip(),
                "categoria": categoria,
                "unidadeProducao": unidade,
                "precoCusto": float(ws.cell(row=r, column=COL_PRECO_CUSTO).value or 0),
                "precoVenda": float(ws.cell(row=r, column=COL_PRECO_VENDA).value or 0),
                "statusVenda": ws.cell(row=r, column=COL_STATUS_VENDA).value or "Ativo",
                # Campos abaixo NAO existem na planilha de origem -- entram com
                # valor padrao seguro e devem ser revisados no Cadastro de Produtos.
                "ativoNaProducao": (ws.cell(row=r, column=COL_STATUS_VENDA).value == "Ativo"),
                "pesoMedioUnitarioGramas": None,
            }
        )
        relatorio.importados += 1

    with open(caminho_saida, "w", encoding="utf-8") as f:
        json.dump(produtos, f, ensure_ascii=False, indent=2)

    return relatorio


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(f"Uso: python3 {sys.argv[0]} <entrada.xlsx> <saida.json>", file=sys.stderr)
        sys.exit(1)

    rel = importar(sys.argv[1], sys.argv[2])
    print("--- Relatorio de Importacao ---")
    print(f"Linhas processadas:      {rel.total_linhas}")
    print(f"Produtos importados:     {rel.importados}")
    print(f"Sem categoria original:  {rel.sem_categoria}  (marcados como '{CATEGORIA_PADRAO}')")
    print(f"Codigos PDV duplicados:  {len(rel.codigos_pdv_duplicados)} {rel.codigos_pdv_duplicados[:10]}")
    print(f"Unidades nao reconhecidas: {len(rel.unidades_nao_reconhecidas)}")
    for aviso in rel.unidades_nao_reconhecidas[:10]:
        print(f"  - {aviso}")
